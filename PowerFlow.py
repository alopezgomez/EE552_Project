# EE 454 Project in Python
# File PowerFlow.py

import math     # Math functions sin, cos
import cmath    # Complex math function conj, rect
import openpyxl # Methods to read and write xlsx files
import numpy    # Methods for linear algebra
# import xlsxwriter

#############################
#
# Start of Main procedure
#
#############################

###################
#
# Constants
#
###################
MVABASE = 100.0 # Sbase
NITER = 20      # Maximum number of iterations (actually does one less)
EPS = 0.001     # Convergence tolerance, per unit MW and MVAR

#############################
#
# Read in bus and line data from Excel spreadsheet
#
#############################

# Open the Excel file with the project data
# data_only means we do not get formulas (good!)
filepath = "system_basecase.xlsx"

"""
import pandas as pd


df = pd.read_excel(filepath,sheet_name=None) #this requires the xlrd library

#pandas dataframes
print(df['BusData'])
print(df['LineData'])

"""
wb = openpyxl.load_workbook(filepath, data_only=True)

# Read in the bus data

# Point at the BusData sheet.
# (This code does not do a any of validation. It expects the sheet names
# to be correct and data to be correctly entered.)
sheet = wb['BusData']

# Make a list of rows
rows = sheet.rows

# Initialize lists.
Pload = []      #Bus load real power, per unit
Qload = []      #Bus load reactive power, per unit
BusType = []    #Bus type: S - Swing, G - Generator (PV), D - Load (PQ)
Pgen = []       #Bus generator real power, per unit
Vset = []       #Bus voltage setpoint, per unit
first = True        # Flag to skip first row

for row in rows:
    if not first: # Skip first row
        Pload.append(float(row[1].value)/MVABASE)
        Qload.append(float(row[2].value)/MVABASE)
        BusType.append(row[3].value)
        Pgen.append(float(row[4].value)/MVABASE)
        Vset.append(float(row[5].value))
    else:
        first = False

# Read in the line data

# Point at the BusData sheet.
sheet = wb['LineData']

# Make a list of rows
rows = sheet.rows

# Initialize lists.
FromBus = []        #Index of from end bus
ToBus = []          #Index of to end bus
R = []              #Series line resistance, per unit
X = []              #Series line reactance, per unit
#Bc = []             #Total line charging, per unit
Pmax = []           #Line flow limit, per unit. 99.99 means no limit
first = True        # Flag to skip first row

for row in rows:
    #print('row',row)
    if not first: # Skip first row
        FromBus.append(row[0].value)
        ToBus.append(row[1].value)
        R.append(float(row[2].value))
        X.append(float(row[3].value))
        #Bc.append(float(row[4].value))
        Pmax.append(float(row[4].value)) ## Keep limit in MVA, only used w/MVA
    else:
        first = False
            
#Done. Note workbook is closed automatically by openpyxl

##############################
#
# Make the admittance matrix (Ybus)
#
##############################

# Initialize to zero
nbus = len(Pload)
# Store Ybus in real and imaginary parts, since G and B appear
# in mismatch and Jacobian equations
Gbus = [[0.] * nbus for i in range(nbus)]
Bbus = [[0.] * nbus for i in range(nbus)]

# For each line
for ln in range (len(FromBus)):
    # Get buses i and j - note, start at zero
    i = FromBus[ln] - 1
    j = ToBus[ln] - 1
    # Find series admittance Y
    Y = 1.0/(R[ln] + 1j * X[ln])
    # Insert in Ybus - diagonals for Yij
    Gbus[i][i] += Y.real
    Bbus[i][i] += Y.imag
    Gbus[j][j] += Y.real
    Bbus[j][j] += Y.imag
    # Off diagonals for Yij
    Gbus[i][j] -= Y.real
    Bbus[i][j] -= Y.imag
    Gbus[j][i] -= Y.real
    Bbus[j][i] -= Y.imag
    #Shunt susceptance Yii (Bc is total, add half at each end)
    #Bbus[i][i] += Bc[ln]/2.0
    #Bbus[j][j] += Bc[ln]/2.0

#Done making Ybus

# Convert Ybus to array for printing
Ybus_mat = numpy.array(Gbus) + numpy.array(Bbus)*1j
# Find known injections
#

Pinjk = [0.0]*nbus
Qinjk = [0.0]*nbus
for i in range(nbus):
    if BusType[i]=='D':
        Pinjk[i] = -Pload[i]
    else:
        Pinjk[i] = Pload[i]
    #Pinjk[i] = Pgen[i] - Pload[i]
    Qinjk[i] = -Qload[i]

####################################
#
# Find bus numbers for Jacobian rows
# The Jacobian has one row for each PQ ('D') or PV ('G') bus
# And then one row for each PQ ('D') bus
#
# This approach differs slightly from the notes because the swing and PV buses
# are simply skipped rather than renumbered, but it's the same idea
#
#####################################

Jbus = []

# Put the PQ and PV buses in the upper partition of the Jacobian
for i in range(nbus):
    if (BusType[i] == 'D' or BusType[i] == 'G'):
        Jbus.append(i)    #Choose to store the bus index, rather than number

# Second loop for lower partition. Include D and G buses
for i in range(nbus):
    if (BusType[i] == 'D'):
        Jbus.append(i)

#####################################
#
# Now for the power flow loop
#
#####################################

#
# Initialize voltage solution
#
# Flat start - keep voltages in bus length, not Jrow
#
Vmag = [1.0] * nbus     # Magnitude of voltage solution, per unit
Theta = [0.0] * nbus    # Angle of voltage solution, radian
#
# update Vmag with Vset for gen buses
for i in range(nbus):
    if (BusType[i] == 'S') or (BusType[i] == 'G'):
        Vmag[i] = Vset[i]

# Initialize convergence records
Pmismax = []   # maximum active power mismatch
Pmismaxbus = []  # bus where the maximum active power mismatch occurs
Qmismax = []  # maximum reactive power mismatch
Qmismaxbus = []   # bus where the maximum reactive power mismatch occurs


#
# Start loop
#
for iter in range(NITER):
    #
    #  Calculate injections Pinjc, Qinjc
    #  Then calculate Jbus ordered Mismatch
    #
    # Set initial injections to zero
    Pinjc = [0.0] * nbus
    Qinjc = [0.0] * nbus

    # Now add line flow terms    
    for k in range(nbus):
        for i in range(nbus):
            Pinjc[k] += Vmag[k] * Vmag[i]                        \
                        *(Gbus[k][i]*math.cos(Theta[k]-Theta[i]) \
                        + Bbus[k][i]*math.sin(Theta[k]-Theta[i]))
            Qinjc[k] += Vmag[k] * Vmag[i]                        \
                        *(Gbus[k][i]*math.sin(Theta[k]-Theta[i]) \
                        - Bbus[k][i]*math.cos(Theta[k]-Theta[i]))


    # Make Jbus ordered Mismatch - this is actually minus [Delta P | Delta Q]^t
    Mismatch = [0.0] * (len(Jbus))
    Pmismax.append(0.0)
    Qmismax.append(0.0)
    Pmismaxbus.append(0)
    Qmismaxbus.append(0)
    for i in range(len(Jbus)):
        if (i < nbus-1):
            Mismatch[i] = Pinjk[Jbus[i]] - Pinjc[Jbus[i]]
            if (abs(Mismatch[i]) > Pmismax[iter]):
                Pmismax[iter] = abs(Mismatch[i])
                Pmismaxbus[iter] = Jbus[i]+1
        else:
            Mismatch[i] = Qinjk[Jbus[i]] - Qinjc[Jbus[i]]
            if (abs(Mismatch[i]) > Qmismax[iter]):
                Qmismax[iter] = abs(Mismatch[i])
                Qmismaxbus[iter] = Jbus[i]+1
        
    # Check for convergence

    mismax = max([abs(numpy.max(Mismatch)), abs(numpy.min(Mismatch))])
##    print('iter %d max mismatch %.8f' % (iter, mismax))
    if (mismax < EPS):
        print('Converged!')
        break

    # Check for excess iterations (goes out of loop with voltage solution
    # used to find mismatch
    if (iter == NITER - 1):
        print('Too many iterations')
        break

    # Calculate Jacobian
    
    # Initialize
    Jac = [[0.0 for i in range(len(Jbus))] for j in range(len(Jbus))]




    # [J11]
    for i in range(nbus-1):
        for j in range(nbus-1):
            ibus = Jbus[i]
            jbus = Jbus[j]
            if ibus == jbus:  # Diagonal term
                Jac[i][j] = -Qinjc[ibus] - (Vmag[ibus]**2)*Bbus[ibus][ibus]
            else:   # Off diagonal
                Jac[i][j] = Vmag[ibus] * Vmag[jbus] \
                 * (Gbus[ibus][jbus] * math.sin(Theta[ibus] - Theta[jbus]) \
                 -  Bbus[ibus][jbus] * math.cos(Theta[ibus] - Theta[jbus]))

    # [J21]
    for i in range(nbus-1, len(Jbus)):
        for j in range(nbus-1):
            ibus = Jbus[i]
            jbus = Jbus[j]
            if ibus == jbus:  # Diagonal term
                Jac[i][j] = Pinjc[ibus] - (Vmag[ibus]**2)*Gbus[ibus][ibus]
            else:   # Off diagonal
                Jac[i][j] = - Vmag[ibus] * Vmag[jbus] \
                 * (Gbus[ibus][jbus] * math.cos(Theta[ibus] - Theta[jbus]) \
                 +  Bbus[ibus][jbus] * math.sin(Theta[ibus] - Theta[jbus]))

    # [J12]
    for i in range(nbus-1):
        for j in range(nbus-1,len(Jbus)):
            ibus = Jbus[i]
            jbus = Jbus[j]
            if ibus == jbus:  # Diagonal term
                Jac[i][j] = (Pinjc[ibus] / Vmag[ibus]) \
                            + Vmag[ibus] * Gbus[ibus][ibus]
            else:   # Off diagonal
                Jac[i][j] = - Vmag[ibus] \
                 * (Gbus[ibus][jbus] * math.cos(Theta[ibus] - Theta[jbus]) \
                 +  Bbus[ibus][jbus] * math.sin(Theta[ibus] - Theta[jbus]))

    # [J22]
    for i in range(nbus-1,len(Jbus)):
        for j in range(nbus-1,len(Jbus)):
            ibus = Jbus[i]
            jbus = Jbus[j]
            if ibus == jbus:  # Diagonal term
                Jac[i][j] = (Qinjc[ibus] / Vmag[ibus]) \
                            - Vmag[ibus] * Bbus[ibus][ibus]
            else:   # Off diagonal
                Jac[i][j] = Vmag[ibus] \
                 * (Gbus[ibus][jbus] * math.sin(Theta[ibus] - Theta[jbus]) \
                 -  Bbus[ibus][jbus] * math.cos(Theta[ibus] - Theta[jbus]))

    if iter == 0:
        test = Jac

    #
    # Find the voltage correction
    #
    deltaV = numpy.linalg.solve(Jac, Mismatch)
    #
    #  Apply voltage correction
    #
    for i in range(len(Jbus)):
        if (i < nbus-1):   # in the Thetas
            ibus = Jbus[i]
            Theta[ibus] += deltaV[i]
        else:           # in the magnitudes
            ibus = Jbus[i]
            Vmag[ibus] += deltaV[i]



######################
#
# End of iteration loop
#
######################
print('Exit from iteration loop')

######################
#
# Print results
#
######################
print('Convergence History')
print('Iter  P Mismatch  P Bus  Q Mismatch  Q Bus')
for i in range(len(Pmismax)):
    print('%3d %12.7f %5d %12.7f %5d' % \
         (i, Pmismax[i], Pmismaxbus[i], Qmismax[i], Qmismaxbus[i]))

print('Bus Results')
print('Bus  Vmag   Theta    PG      PL      Pk')
for i in range(nbus):
    print('%3d %6.3f %6.2f  %6.1f   %6.1f   %6.1f' % (i+1, Vmag[i], Theta[i]*180/math.pi,
                                                       (Pinjc[i]-Pload[i])*MVABASE,Pload[i]*MVABASE,Pinjc[i]*MVABASE))



print('Line MVA')
print('Line  From     To  FromMW    FromMVAr  FromMVA  ToMVA    Limit Flag')
for ln in range(len(FromBus)):
    i = FromBus[ln]
    j = ToBus[ln]
    Vi = cmath.rect(Vmag[i-1], Theta[i-1])
    Vj = cmath.rect(Vmag[j-1], Theta[j-1])
    Iij = ((Vi - Vj) / (R[ln] + 1j*X[ln])) #+ Vi * 1j*Bc[ln]/2.0 #pf includes shunt current
    Sij = Vi * Iij.conjugate() * MVABASE
    Pij = numpy.real(Sij)
    Qij = numpy.imag(Sij)
    Sijmag = abs(Sij)
    j = FromBus[ln]
    i = ToBus[ln]
    Vi = cmath.rect(Vmag[i-1], Theta[i-1])
    Vj = cmath.rect(Vmag[j-1], Theta[j-1])
    Iij = ((Vi - Vj) / (R[ln] + 1j*X[ln])) #+ Vi * 1j*Bc[ln]/2.0  #pf includes shunt current
    Sji = Vi * Iij.conjugate() * MVABASE
    Pji = numpy.real(Sji)
    Qji = numpy.imag(Sji)
    Sjimag = abs(Sji)
    if (Sijmag > Pmax[ln]) or (Sjimag > Pmax[ln]):
        flag = '***'
    else:
        flag = ''
    print('%3d %6d %6d %8.2f %8.2f %8.2f %8.2f %8.1f %s' % \
          (ln+1, FromBus[ln], ToBus[ln], Pij, Qij, Sijmag, Sjimag, Pmax[ln], flag))
    print('%3d %6d %6d %8.2f %8.2f %8.2f %8.2f %8.1f %s' % \
          (ln + 1, ToBus[ln], FromBus[ln], Pji, Qji, Sjimag, Sijmag, Pmax[ln], flag))

    


# save results to files
#
# gfile = r"C:\Users\trisharay\Documents\MSEE\2020-Autumn\EE 454\Test System Code\Gmatrix.xlsx"
# bfile = r"C:\Users\trisharay\Documents\MSEE\2020-Autumn\EE 454\Test System Code\Bmatrix.xlsx"
#
# workbookG = xlsxwriter.Workbook(gfile)
# workbookB = xlsxwriter.Workbook(bfile)
# worksheetG = workbookG.add_worksheet()
# worksheetB = workbookB.add_worksheet()
#
#
# row = 0
#
# for col, data in enumerate(Ybus_mat):
#     worksheetG.write_column(row, col, data.real)
#     worksheetB.write_column(row,col, data.imag)
#
# workbookG.close()
# workbookB.close()

#
# import csv
# enumerate(Ybus_mat)
# with open('some.csv', 'w', newline='') as f:
#     writer = csv.writer(f)
#     for row in Ybus_mat:
#         writer.writerows(row)
#
# #

# busresultsfile = r"C:\Users\trisharay\Documents\MSEE\2020-Autumn\EE 454\Test System Code\BusResults.xlsx"

# bookBusRes = xlsxwriter.Workbook(busresultsfile)
# sheetBusRes = bookBusRes.add_worksheet()

# for col in range(nbus):
#     data = (i + 1, Vmag[i], Theta[i] * 180 / math.pi, (Pinjc[i] - Pload[i]) * MVABASE, Pload[i] * MVABASE, Pinjc[i] * MVABASE)
#     print('data',data)
#     sheetBusRes.write_column(row, col, [])

#     for i in range(nbus):
#         print('%3d %6.3f %6.2f  %6.1f   %6.1f   %6.1f' % (i + 1, Vmag[i], Theta[i] * 180 / math.pi,
#                                                           (Pinjc[i] - Pload[i]) * MVABASE, Pload[i] * MVABASE,
#                                                           Pinjc[i] * MVABASE))